import os
import pandas as pd
import psycopg2
import datetime
from openpyxl import load_workbook

# --- 1. Database connection from GitHub Secrets ---
conn = psycopg2.connect(
    host=os.environ["DB_HOST"],
    port=os.environ["DB_PORT"],
    dbname=os.environ["DB_NAME"],
    user=os.environ["DB_USER"],
    password=os.environ["DB_PASSWORD"],
    sslmode="require"
)

# --- 2. Fetch data ---
df_kigyou = pd.read_sql("SELECT * FROM production.v_kigyou_csv;", conn)
df_user = pd.read_sql("SELECT * FROM production.v_user_csv;", conn)
conn.close()

# --- 3. Set output Excel file inside repo ---
today = datetime.datetime.now().strftime("%m%d")
src_file = "template/企業一覧とユーザー一覧.xlsx"  # Template file checked into repo
dest_file = f"output/企業一覧とユーザー一覧_{today}.xlsx"

# Make sure output folder exists
os.makedirs("output", exist_ok=True)

# Copy template
import shutil
shutil.copy(src_file, dest_file)

# --- 4. Update Sheets ---
wb = load_workbook(dest_file)

# Sheet 1: 企業一覧
ws1 = wb["企業一覧"]
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    for cell in row:
        cell.value = None
for i, row in enumerate(df_kigyou.values, start=2):
    for j, value in enumerate(row, start=1):
        ws1.cell(row=i, column=j, value=value)

# Sheet 2: ユーザー一覧
ws2 = wb["ユーザー一覧"]
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for cell in row:
        cell.value = None
for i, row in enumerate(df_user.values, start=2):
    for j, value in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=value)

wb.save(dest_file)
print(f"Export complete: {dest_file}")